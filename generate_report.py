"""
Cron job: Runs every morning at 9AM.
Generates a NEW Excel sheet for each day's sent emails.
File: reports/sent_emails_YYYY-MM-DD.xlsx
Also updates a master report: reports/all_sent_emails.xlsx
"""

import json
import os
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))

SENT_LOG_FILE = os.path.join(BASE_DIR, "sent_log.json")
REPORTS_DIR = os.path.join(BASE_DIR, "reports")
MASTER_REPORT = os.path.join(REPORTS_DIR, "all_sent_emails.xlsx")

HEADERS = ["S.No", "Email", "Name", "Company", "Title", "Source", "Sent At", "Status"]


def load_json(filepath):
    if os.path.exists(filepath):
        with open(filepath, "r") as f:
            return json.load(f)
    return []


def create_sheet(filepath, entries):
    """Create an Excel file with given entries"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sent Emails"
    ws.append(HEADERS)

    # Bold header
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)

    for i, entry in enumerate(entries, 1):
        ws.append([
            i,
            entry.get("email", ""),
            entry.get("name", ""),
            entry.get("company", ""),
            entry.get("title", ""),
            entry.get("source", ""),
            entry.get("sent_at", ""),
            entry.get("status", ""),
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(filepath)
    return len(entries)


def update_master_report(all_sent):
    """Update/recreate the master report with ALL sent emails"""
    create_sheet(MASTER_REPORT, all_sent)


def main():
    os.makedirs(REPORTS_DIR, exist_ok=True)

    sent_log = load_json(SENT_LOG_FILE)
    all_sent = [e for e in sent_log if e.get("status") == "sent"]

    if not all_sent:
        print("No sent emails found in log.")
        return

    # Yesterday's date (report for previous day's sends)
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    today = datetime.now().strftime("%Y-%m-%d")

    # Filter emails sent yesterday
    yesterday_sent = [
        e for e in all_sent
        if e.get("sent_at", "").startswith(yesterday)
    ]

    # Also check today (in case cron runs right after midnight sends)
    today_sent = [
        e for e in all_sent
        if e.get("sent_at", "").startswith(today)
    ]

    # Generate daily report for yesterday
    if yesterday_sent:
        daily_file = os.path.join(REPORTS_DIR, f"sent_emails_{yesterday}.xlsx")
        count = create_sheet(daily_file, yesterday_sent)
        print(f"Daily report: {daily_file} ({count} emails)")
    else:
        print(f"No emails sent on {yesterday}")

    # Generate daily report for today if any
    if today_sent:
        daily_file = os.path.join(REPORTS_DIR, f"sent_emails_{today}.xlsx")
        count = create_sheet(daily_file, today_sent)
        print(f"Today's report: {daily_file} ({count} emails)")

    # Update master report
    update_master_report(all_sent)
    print(f"Master report: {MASTER_REPORT} ({len(all_sent)} total emails)")


if __name__ == "__main__":
    main()
